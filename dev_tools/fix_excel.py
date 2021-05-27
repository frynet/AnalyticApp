# -*- coding: utf-8 -*-
from typing import Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell, MergedCell


def fix_excel_file(file_name):
    rb = load_workbook(file_name)
    wb = Workbook()
    for sheet_name in rb.sheetnames[1:]:
        header = False
        value = None
        sheet: Worksheet = rb[sheet_name]
        w_sheet: Worksheet = wb.create_sheet(sheet_name)
        for row_r_i, row in enumerate(sheet.rows):
            row: Tuple[Cell]
            is_all_str = True
            true_values = []
            for col_i, cell in enumerate(row):
                if cell.value is not None:
                    for m_cell_range in sheet.merged_cells.ranges:
                        if m_cell_range.min_col != m_cell_range.max_col \
                                and cell.coordinate in m_cell_range \
                                and cell.data_type == "s":
                            m_cell_range.shift(row_shift=1)
                            true_values.extend([sheet.cell(*cell).value for cell in m_cell_range.cells])
                            break
                    else:
                        true_values.append(cell.value)
                        if cell.data_type != 's':
                            is_all_str = False
            if true_values:
                print(true_values)
                if not is_all_str:

                    if len(true_values) == 1:
                        value = true_values[0]
                    elif len(true_values) > 1:
                        w_sheet.append([value] + true_values)
                elif not header:
                    w_sheet.append(["Value"] + true_values)
                    header = True

    del wb[wb.sheetnames[0]]
    wb.save("./../data/lake11.xlsx")


if __name__ == '__main__':
    fix_excel_file("./../data/lake1.xlsx")
